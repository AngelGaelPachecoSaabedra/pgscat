"""
clinical_genes_builder.py
=========================
Builds the authoritative clinical genes dataset by combining:

  1. LOCAL Excel  — path configured via --excel flag (see usage below)
     600 AR genes curated from GenCC with Definitive/Strong/Moderate evidence.
     Source label: "LocalExcel_GenCC_AR"

  2. GenCC CSV    — resources/clinical_genes/raw/gencc_submissions.csv
     Full GenCC public submissions (all MOI, all evidence tiers).
     Filtered to: Autosomal recessive + evidence ≥ Moderate.
     Source label: "GenCC"

  3. ClinVar gene-condition (optional)
     resources/clinical_genes/raw/clinvar_gene_condition.txt
     Used only to cross-validate; genes not added from ClinVar alone
     unless they are also in GenCC or LocalExcel.
     Source label: "ClinVar"

Output:
  resources/clinical_genes/clinical_genes_v1.json
  resources/clinical_genes/clinical_genes_v1.csv
  resources/clinical_genes/validation_report.json

Usage (standalone):
  python3 src/services/clinical_genes_builder.py

  Optional flags:
    --excel PATH   Override Excel file location
    --raw-dir DIR  Override raw downloads directory
    --out-dir DIR  Override output directory
    --annotations-dir DIR  Compare against annotated genes
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("clinical_genes_builder")

# ── Default paths ─────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent  # src/services/
_APP_ROOT = _HERE.parent.parent  # app/

EXCEL_PATH_DEFAULT    = _APP_ROOT / "resources" / "clinical_genes" / "genes_AR_600_GenCC.xlsx"
RAW_DIR_DEFAULT       = _APP_ROOT / "resources" / "clinical_genes" / "raw"
OUT_DIR_DEFAULT       = _APP_ROOT / "resources" / "clinical_genes"
ANNOTATIONS_DIR_DEFAULT = Path(os.environ.get("APP_ANNOTATIONS_DIR", "/annotations"))

# GenCC evidence tiers considered high-confidence (in order of decreasing strength)
_GENCC_HIGH_CONFIDENCE = {
    "definitive", "strong", "moderate",
}

# GenCC MOI values that map to autosomal recessive
_AR_MOI_KEYWORDS = {"autosomal recessive", "ar"}


# ── Normalisation helpers ─────────────────────────────────────────────────────

def _normalise_symbol(raw: str) -> Optional[str]:
    """Return uppercase stripped gene symbol, or None if empty/invalid."""
    if not raw or not isinstance(raw, str):
        return None
    s = raw.strip().upper()
    if not s or s in {".", "-", "NA", "NAN", "N/A", ""}:
        return None
    return s


def _normalise_evidence(raw: str) -> str:
    """Return lowercase stripped evidence tier."""
    if not raw or not isinstance(raw, str):
        return "unknown"
    return raw.strip().lower()


def _is_ar(moi: str) -> bool:
    """Return True if the MOI string represents autosomal recessive."""
    if not moi or not isinstance(moi, str):
        return False
    m = moi.strip().lower()
    return any(k in m for k in _AR_MOI_KEYWORDS)


# ── Source loaders ────────────────────────────────────────────────────────────

def load_local_excel(excel_path: Path) -> List[Dict[str, Any]]:
    """
    Load genes from the local Excel file.

    Expected sheet: 'AR_600_genes'
    Expected columns: Gene, HGNC ID (CURIE), Enfermedad (MONDO), Disease CURIE,
                      Referencia original, Nombre original, Herencia (MOI),
                      Evidencia (GenCC), Submitter, PMIDs, ...

    Returns list of dicts with normalised fields.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install it: pip install openpyxl")
        return []

    if not excel_path.exists():
        logger.warning("Local Excel not found: %s", excel_path)
        return []

    logger.info("Loading local Excel: %s", excel_path)
    genes: List[Dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        # Try the expected sheet name first
        sheet_name = "AR_600_genes" if "AR_600_genes" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        headers: Optional[List[str]] = None
        col_gene = col_moi = col_evidence = col_disease = col_hgnc = None
        duplicates_seen: Set[str] = set()

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # Header row
                headers = [str(h).strip() if h else "" for h in row]
                h_lower = [h.lower() for h in headers]

                # Find gene symbol column
                for candidate in ("gene", "gene_symbol", "symbol", "hgnc_symbol"):
                    if candidate in h_lower:
                        col_gene = h_lower.index(candidate)
                        break
                if col_gene is None:
                    # Try first column
                    col_gene = 0
                    logger.warning("Could not detect gene column; using column 0 (%s)", headers[0])

                # Find MOI column
                for candidate in ("herencia (moi)", "herencia", "moi", "inheritance", "mode of inheritance"):
                    if candidate in h_lower:
                        col_moi = h_lower.index(candidate)
                        break

                # Find evidence column
                for candidate in ("evidencia (gencc)", "evidencia", "evidence", "classification", "gencc_classification"):
                    if candidate in h_lower:
                        col_evidence = h_lower.index(candidate)
                        break

                # Find disease column
                for candidate in ("enfermedad (mondo)", "disease", "condition", "enfermedad", "nombre original"):
                    if candidate in h_lower:
                        col_disease = h_lower.index(candidate)
                        break

                # Find HGNC column
                for candidate in ("hgnc id (curie)", "hgnc_id", "hgnc", "hgnc id"):
                    if candidate in h_lower:
                        col_hgnc = h_lower.index(candidate)
                        break

                logger.info(
                    "Excel columns: gene=%s, moi=%s, evidence=%s, disease=%s",
                    col_gene, col_moi, col_evidence, col_disease
                )
                continue

            # Data row
            if not row or all(v is None for v in row):
                continue

            raw_gene = row[col_gene] if col_gene is not None and len(row) > col_gene else None
            symbol = _normalise_symbol(str(raw_gene) if raw_gene is not None else "")
            if not symbol:
                continue

            if symbol in duplicates_seen:
                logger.debug("Excel duplicate skipped: %s", symbol)
                continue
            duplicates_seen.add(symbol)

            moi = str(row[col_moi]).strip() if col_moi is not None and len(row) > col_moi and row[col_moi] else "Autosomal recessive"
            evidence = str(row[col_evidence]).strip() if col_evidence is not None and len(row) > col_evidence and row[col_evidence] else "unknown"
            disease = str(row[col_disease]).strip() if col_disease is not None and len(row) > col_disease and row[col_disease] else ""
            hgnc_id = str(row[col_hgnc]).strip() if col_hgnc is not None and len(row) > col_hgnc and row[col_hgnc] else ""

            genes.append({
                "gene_symbol": symbol,
                "sources": ["LocalExcel_GenCC_AR"],
                "moi": moi,
                "evidence": _normalise_evidence(evidence),
                "disease": disease,
                "hgnc_id": hgnc_id,
            })

        wb.close()
        logger.info("Excel: loaded %d unique genes (sheet: %s)", len(genes), sheet_name)

    except Exception as exc:
        logger.error("Error loading Excel %s: %s", excel_path, exc)

    return genes


def load_gencc_csv(csv_path: Path) -> List[Dict[str, Any]]:
    """
    Load AR genes from GenCC submissions CSV.

    Filters: MOI = autosomal recessive, classification ≥ Moderate.
    Returns list of dicts. Keeps only one row per gene (highest confidence).
    """
    if not csv_path.exists():
        logger.warning("GenCC CSV not found: %s", csv_path)
        return []

    logger.info("Loading GenCC CSV: %s", csv_path)

    # Evidence tier ordering (higher index = higher confidence)
    _TIER_RANK = {
        "definitive": 5, "strong": 4, "moderate": 3,
        "limited": 2, "supportive": 1, "disputed": 0,
        "refuted": 0, "no known disease relationship": 0,
    }

    # Per-gene best entry tracking
    gene_best: Dict[str, Dict[str, Any]] = {}

    try:
        with open(csv_path, encoding="utf-8", errors="replace", newline="") as fh:
            reader = csv.DictReader(fh)
            headers = reader.fieldnames or []
            h_lower = {h.lower().strip(): h for h in headers}

            # Detect columns (GenCC CSV columns may vary slightly)
            col_gene = None
            for candidate in ("gene_symbol", "gene symbol", "hgnc_gene_symbol", "symbol"):
                if candidate in h_lower:
                    col_gene = h_lower[candidate]
                    break
            if col_gene is None:
                # Look for any column with 'gene' and 'symbol'
                for h in headers:
                    if "gene" in h.lower() and "symbol" in h.lower():
                        col_gene = h
                        break
            if col_gene is None:
                logger.warning("Could not find gene symbol column in GenCC CSV. Headers: %s", headers[:10])
                return []

            col_moi = None
            for candidate in ("moi_title", "moi", "mode_of_inheritance", "inheritance"):
                if candidate in h_lower:
                    col_moi = h_lower[candidate]
                    break

            col_class = None
            for candidate in ("classification_title", "classification", "evidence", "gencc_classification"):
                if candidate in h_lower:
                    col_class = h_lower[candidate]
                    break

            col_disease = None
            for candidate in ("disease_title", "disease", "condition_title", "condition"):
                if candidate in h_lower:
                    col_disease = h_lower[candidate]
                    break

            col_submitter = None
            for candidate in ("submitter_title", "submitter", "source"):
                if candidate in h_lower:
                    col_submitter = h_lower[candidate]
                    break

            logger.info(
                "GenCC CSV columns detected: gene=%s, moi=%s, class=%s",
                col_gene, col_moi, col_class
            )

            n_total = n_ar = n_hc = 0
            for row in reader:
                n_total += 1
                raw_symbol = row.get(col_gene, "") if col_gene else ""
                symbol = _normalise_symbol(raw_symbol)
                if not symbol:
                    continue

                moi_raw = row.get(col_moi, "") if col_moi else ""
                if not _is_ar(moi_raw):
                    continue
                n_ar += 1

                classification = _normalise_evidence(row.get(col_class, "") if col_class else "")
                tier = _TIER_RANK.get(classification, -1)
                if tier < _TIER_RANK.get("moderate", 3):
                    continue
                n_hc += 1

                disease = row.get(col_disease, "").strip() if col_disease else ""
                submitter = row.get(col_submitter, "").strip() if col_submitter else ""

                if symbol not in gene_best or tier > _TIER_RANK.get(gene_best[symbol]["evidence"], -1):
                    gene_best[symbol] = {
                        "gene_symbol": symbol,
                        "sources": ["GenCC"],
                        "moi": moi_raw.strip(),
                        "evidence": classification,
                        "disease": disease,
                        "submitter": submitter,
                        "hgnc_id": "",
                    }

            logger.info(
                "GenCC CSV: %d total rows → %d AR → %d high-confidence AR genes",
                n_total, n_ar, n_hc
            )

    except Exception as exc:
        logger.error("Error reading GenCC CSV %s: %s", csv_path, exc)
        return []

    return list(gene_best.values())


def load_clinvar_gene_condition(txt_path: Path) -> Set[str]:
    """
    Load gene symbols from ClinVar gene-condition mapping.
    Returns only the set of symbols (used for cross-validation, not as a primary source).

    File format: tab-separated, columns include GeneSymbol.
    """
    if not txt_path.exists():
        return set()

    logger.info("Loading ClinVar gene list from: %s", txt_path)
    symbols: Set[str] = set()

    try:
        with open(txt_path, encoding="utf-8", errors="replace") as fh:
            reader = csv.reader(fh, delimiter="\t")
            headers = next(reader, [])
            h_lower = [h.lower().strip() for h in headers]

            col_gene = None
            for candidate in ("genesymbol", "gene_symbol", "gene symbol", "symbol"):
                if candidate in h_lower:
                    col_gene = h_lower.index(candidate)
                    break
            if col_gene is None:
                col_gene = 0

            for row in reader:
                if not row or len(row) <= col_gene:
                    continue
                sym = _normalise_symbol(row[col_gene])
                if sym:
                    symbols.add(sym)

        logger.info("ClinVar: %d unique gene symbols loaded", len(symbols))
    except Exception as exc:
        logger.warning("Cannot read ClinVar gene list %s: %s", txt_path, exc)

    return symbols


def load_annotated_genes(annotations_dir: Path) -> Set[str]:
    """
    Collect all gene symbols from annotation parquet/tsv files.
    Used in validation report: which clinical genes are present in annotations?
    """
    genes: Set[str] = set()
    if not annotations_dir.exists():
        return genes

    try:
        import pandas as pd
    except ImportError:
        return genes

    for subdir in sorted(annotations_dir.iterdir()):
        if not subdir.is_dir():
            continue
        parquet = subdir / f"{subdir.name}_variants_annotated.parquet"
        tsv = subdir / f"{subdir.name}_variants_annotated.tsv.gz"

        try:
            if parquet.exists():
                try:
                    import pyarrow.parquet as pq
                    df = pq.read_table(str(parquet), columns=["gene_name"]).to_pandas()
                except Exception:
                    df = pd.read_parquet(str(parquet), columns=["gene_name"])
            elif tsv.exists():
                df = pd.read_csv(tsv, sep="\t", usecols=["gene_name"],
                                 compression="infer", low_memory=False)
            else:
                continue

            if "gene_name" in df.columns:
                for g in df["gene_name"].dropna().unique():
                    sym = _normalise_symbol(str(g))
                    if sym:
                        genes.add(sym)
        except Exception as exc:
            logger.debug("Could not read %s: %s", subdir.name, exc)

    logger.info("Annotated genes found in parquets: %d", len(genes))
    return genes


# ── Merger ────────────────────────────────────────────────────────────────────

def merge_sources(
    excel_genes: List[Dict[str, Any]],
    gencc_genes: List[Dict[str, Any]],
    clinvar_symbols: Set[str],
) -> List[Dict[str, Any]]:
    """
    Merge gene lists from all sources into a single deduplicated dataset.

    Strategy:
      - Start with all genes from LocalExcel_GenCC_AR
      - Add GenCC genes not already present
      - Mark genes also present in ClinVar
      - Assign confidence based on source overlap and evidence tier

    Confidence tiers:
      high    — in LocalExcel AND GenCC with Definitive/Strong
      medium  — in GenCC with Definitive/Strong, or LocalExcel + ClinVar
      standard — GenCC Moderate, or single source
    """
    # Map: gene_symbol → merged entry
    merged: Dict[str, Dict[str, Any]] = {}

    # ── Process Excel genes (primary: already curated AR set) ─────────────────
    for entry in excel_genes:
        sym = entry["gene_symbol"]
        merged[sym] = {
            "gene_symbol": sym,
            "sources": list(entry.get("sources", ["LocalExcel_GenCC_AR"])),
            "moi": entry.get("moi", "Autosomal recessive"),
            "evidence": entry.get("evidence", "unknown"),
            "disease": entry.get("disease", ""),
            "hgnc_id": entry.get("hgnc_id", ""),
            "in_clinvar": sym in clinvar_symbols,
        }

    # ── Merge GenCC genes ─────────────────────────────────────────────────────
    for entry in gencc_genes:
        sym = entry["gene_symbol"]
        if sym in merged:
            # Gene already present; merge sources and upgrade evidence if better
            existing = merged[sym]
            if "GenCC" not in existing["sources"]:
                existing["sources"].append("GenCC")
            # Use better evidence tier
            _TIER = {"definitive": 5, "strong": 4, "moderate": 3, "limited": 2, "unknown": 0}
            if _TIER.get(entry.get("evidence", ""), 0) > _TIER.get(existing.get("evidence", ""), 0):
                existing["evidence"] = entry["evidence"]
            # Keep disease from GenCC if Excel has none
            if not existing.get("disease") and entry.get("disease"):
                existing["disease"] = entry["disease"]
        else:
            merged[sym] = {
                "gene_symbol": sym,
                "sources": ["GenCC"],
                "moi": entry.get("moi", "Autosomal recessive"),
                "evidence": entry.get("evidence", "unknown"),
                "disease": entry.get("disease", ""),
                "hgnc_id": entry.get("hgnc_id", ""),
                "in_clinvar": sym in clinvar_symbols,
            }

    # ── Mark ClinVar presence ─────────────────────────────────────────────────
    for sym, entry in merged.items():
        if sym in clinvar_symbols and "ClinVar" not in entry["sources"]:
            entry["in_clinvar"] = True

    # ── Assign confidence ─────────────────────────────────────────────────────
    _TIER = {"definitive": 5, "strong": 4, "moderate": 3, "limited": 2, "unknown": 0}
    for entry in merged.values():
        sources = set(entry["sources"])
        ev_rank = _TIER.get(entry.get("evidence", ""), 0)
        in_local = "LocalExcel_GenCC_AR" in sources
        in_gencc = "GenCC" in sources
        in_clinvar = entry.get("in_clinvar", False)

        if in_local and in_gencc and ev_rank >= 4:
            conf = "high"
        elif in_local and ev_rank >= 4:
            conf = "high"
        elif (in_local or in_gencc) and ev_rank >= 3:
            conf = "medium"
        elif in_local and in_clinvar:
            conf = "medium"
        else:
            conf = "standard"

        entry["confidence"] = conf

    # Return sorted by gene symbol
    result = sorted(merged.values(), key=lambda g: g["gene_symbol"])
    logger.info("Merged: %d unique clinical genes", len(result))
    return result


# ── Validation report ─────────────────────────────────────────────────────────

def build_validation_report(
    merged: List[Dict[str, Any]],
    excel_genes: List[Dict[str, Any]],
    gencc_genes: List[Dict[str, Any]],
    clinvar_symbols: Set[str],
    annotated_genes: Set[str],
) -> Dict[str, Any]:
    """Build a validation/overlap report across all sources."""
    merged_syms  = {g["gene_symbol"] for g in merged}
    excel_syms   = {g["gene_symbol"] for g in excel_genes}
    gencc_syms   = {g["gene_symbol"] for g in gencc_genes}
    clinvar_syms = clinvar_symbols
    ann_syms     = annotated_genes

    # Confidence breakdown
    conf_counts = defaultdict(int)
    for g in merged:
        conf_counts[g["confidence"]] += 1

    # Source overlap
    in_excel_and_gencc = excel_syms & gencc_syms
    in_excel_only      = excel_syms - gencc_syms
    in_gencc_only      = gencc_syms - excel_syms

    # Annotation overlap
    clinical_in_annotations = merged_syms & ann_syms if ann_syms else set()
    clinical_missing_from_annotations = merged_syms - ann_syms if ann_syms else set()

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "totals": {
            "n_merged": len(merged),
            "n_excel": len(excel_syms),
            "n_gencc": len(gencc_syms),
            "n_clinvar": len(clinvar_syms),
            "n_annotated_genes": len(ann_syms),
        },
        "confidence_breakdown": dict(conf_counts),
        "source_overlap": {
            "in_excel_and_gencc": len(in_excel_and_gencc),
            "in_excel_only": len(in_excel_only),
            "in_gencc_only": len(in_gencc_only),
        },
        "annotation_overlap": {
            "clinical_genes_in_annotations": len(clinical_in_annotations),
            "clinical_genes_missing_from_annotations": len(clinical_missing_from_annotations),
            "pct_covered": (
                round(len(clinical_in_annotations) / len(merged) * 100, 1)
                if merged else 0.0
            ),
        },
        "genes_only_in_excel": sorted(in_excel_only)[:50],
        "genes_only_in_gencc": sorted(in_gencc_only)[:50],
        "clinical_genes_in_annotations": sorted(clinical_in_annotations)[:100],
        "clinical_genes_missing_from_annotations": sorted(clinical_missing_from_annotations)[:100],
    }
    return report


# ── Writers ───────────────────────────────────────────────────────────────────

def write_json(genes: List[Dict[str, Any]], out_path: Path) -> None:
    """Write clinical genes to JSON."""
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "n_genes": len(genes),
        "description": (
            "Clinical genes dataset combining LocalExcel_GenCC_AR (600 AR genes curated "
            "from GenCC) and GenCC public submissions (AR, evidence ≥ Moderate)."
        ),
        "genes": genes,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Wrote JSON: %s (%d genes)", out_path, len(genes))


def write_csv(genes: List[Dict[str, Any]], out_path: Path) -> None:
    """Write clinical genes to CSV."""
    if not genes:
        return
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["gene_symbol", "sources", "confidence", "moi", "evidence", "disease", "hgnc_id", "in_clinvar"]
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for g in genes:
            row = dict(g)
            row["sources"] = ";".join(g.get("sources", []))
            writer.writerow(row)
    logger.info("Wrote CSV: %s (%d genes)", out_path, len(genes))


# ── Main ──────────────────────────────────────────────────────────────────────

def build(
    excel_path: Path = EXCEL_PATH_DEFAULT,
    raw_dir: Path = RAW_DIR_DEFAULT,
    out_dir: Path = OUT_DIR_DEFAULT,
    annotations_dir: Path = ANNOTATIONS_DIR_DEFAULT,
) -> Dict[str, Any]:
    """
    Run the full clinical genes build pipeline.
    Returns the validation report dict.
    """
    logger.info("=== Clinical Genes Builder ===")
    logger.info("  Excel       : %s", excel_path)
    logger.info("  Raw dir     : %s", raw_dir)
    logger.info("  Output dir  : %s", out_dir)
    logger.info("  Annotations : %s", annotations_dir)

    # ── Load sources ──────────────────────────────────────────────────────────
    excel_genes    = load_local_excel(excel_path)
    gencc_genes    = load_gencc_csv(raw_dir / "gencc_submissions.csv")
    clinvar_syms   = load_clinvar_gene_condition(raw_dir / "clinvar_gene_condition.txt")
    annotated_syms = load_annotated_genes(annotations_dir)

    if not excel_genes and not gencc_genes:
        logger.error(
            "Neither local Excel nor GenCC CSV could be loaded. "
            "Run scripts/download_clinical_genes.sh first."
        )
        sys.exit(1)

    # ── Merge ─────────────────────────────────────────────────────────────────
    merged = merge_sources(excel_genes, gencc_genes, clinvar_syms)

    # ── Write outputs ─────────────────────────────────────────────────────────
    out_dir.mkdir(parents=True, exist_ok=True)
    write_json(merged, out_dir / "clinical_genes_v1.json")
    write_csv(merged,  out_dir / "clinical_genes_v1.csv")

    # ── Validation report ─────────────────────────────────────────────────────
    report = build_validation_report(
        merged, excel_genes, gencc_genes, clinvar_syms, annotated_syms
    )
    report_path = out_dir / "validation_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Wrote validation report: %s", report_path)

    # ── Summary ───────────────────────────────────────────────────────────────
    logger.info("")
    logger.info("=== Build Summary ===")
    logger.info("  Local Excel genes : %d", len(excel_genes))
    logger.info("  GenCC AR genes    : %d", len(gencc_genes))
    logger.info("  ClinVar symbols   : %d", len(clinvar_syms))
    logger.info("  Merged total      : %d", len(merged))
    logger.info("  Confidence:")
    for c, n in sorted(report["confidence_breakdown"].items()):
        logger.info("    %-10s : %d", c, n)
    logger.info("  Annotation overlap: %d / %d (%.1f%%)",
        report["annotation_overlap"]["clinical_genes_in_annotations"],
        len(merged),
        report["annotation_overlap"]["pct_covered"],
    )

    return report


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build clinical genes dataset")
    parser.add_argument("--excel", type=Path, default=EXCEL_PATH_DEFAULT,
                        help="Path to local Excel file")
    parser.add_argument("--raw-dir", type=Path, default=RAW_DIR_DEFAULT,
                        help="Directory with downloaded raw files")
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR_DEFAULT,
                        help="Output directory for JSON/CSV")
    parser.add_argument("--annotations-dir", type=Path, default=ANNOTATIONS_DIR_DEFAULT,
                        help="Annotations directory for overlap validation")
    args = parser.parse_args()

    build(
        excel_path=args.excel,
        raw_dir=args.raw_dir,
        out_dir=args.out_dir,
        annotations_dir=args.annotations_dir,
    )
